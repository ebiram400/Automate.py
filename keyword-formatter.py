import openpyxl
import json
import re

# بارگذاری فایل اکسل
file_path = "Book1.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# تنظیم ستون‌های مورد نظر
columns = {
    "SEO_Keywords": 1,
    "meta": 2
}

# حلقه برای پردازش ردیف‌ها
for i in range(1, ws.max_row + 1):
    sentence = ws.cell(row=i, column=columns["SEO_Keywords"]).value
    if not sentence:
        continue
    li = sentence.split(',')
    target = []
    for j in li:
        j = re.sub(r"\(.*?\)","",j).strip()
        obj = {"keyword":j,"score":0}
        target.append(obj)
    target = json.dumps(target,ensure_ascii=False)
    ws.cell(row=i, column=columns["meta"]).value = target
    wb.save(file_path)

print("************The End***********")