import os
import openpyxl

listPic = os.listdir("E:pictures")
listNamePic = [name.removesuffix(".jpg") for name in listPic]

# بارگذاری فایل اکسل
file_path = "E:product1.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

columns = {
    "شناسه": 1,
    "نوع": 2,
    "شناسه محصول": 3,
    "نام":4,
    "منتشر شده":5,
    "ویژه است؟":6,
    "قابلیت نمایش در کاتالوگ":7,
    "توضیح کوتاه":8,
    "توضیحات":9,
    "تاریخ شروع حراجی":10,
    "تاریخ پایان حراجی":11,
    "وضعیت مالیات":12,
    "کلاس مالیاتی":13,
    "در انبار؟":14,
    "انبار":15,
    "میزان کمبود در انبار":16,
    "میزان کمبود در انبار":17,
    "پیش خرید فعال است؟":18,
    "فروش به صورت تکی؟":19,
    "وزن (کیلوگرم)":20,
    "درازا (سانتیمتر)":21,
    "پهنا (سانتیمتر)":22,
    "بلندا (سانتیمتر)":23,
    "اجازه ارسال نقد و بررسی کاربر امکان پذیر است؟":24,
    "یادداشت خرید":25,
    "قیمت فروش فوق‌العاده":26,
    "قیمت اصلی":27,
    "دسته‌ها":28,
    "برچسب‌ها":29,
    "کلاس ارسال":30,
    "تصاویر":31
}

baseUrl = "https://tabatabaeiabzarmarket.ir/wp-content/uploads/2024/12/"

for i in range(1,ws.max_row + 1):
    namePic=ws.cell(row=i, column=columns["نام"]).value
    namePic = namePic.replace('*','').replace('/','')
    namePic = namePic.replace(" ","_")
    if namePic in listNamePic:
        urlNewPic = baseUrl + namePic +".jpg"
        ws.cell(row=i, column=columns["تصاویر"]).value = urlNewPic
        print("change "+urlNewPic +" ...")

# ذخیره تغییرات بعد از هر به‌روزرسانی
wb.save(file_path)
print("************the end************")