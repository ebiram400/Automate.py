import openpyxl
import google.generativeai as genai
import json
import time

# بارگذاری فایل اکسل
file_path = "product.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# تنظیم ستون‌های مورد نظر
columns = {
    "Name": 1,  # فرض: ستون اول برای نام محصول
    "Product_Description": 2,  # توضیحات محصول
    "Short_Description": 3,  # توضیح مختصر
    "SEO_Keywords": 4,  # کلمات کلیدی
    "Tags": 5  # برچسب‌ها
}

# حلقه برای پردازش ردیف‌ها
for i in range(849, ws.max_row + 1):  # فرض: پردازش از ردیف ... تا 913
    print("generate for", ws.cell(row=i, column=columns["Name"]).value, "...")

    # پیکربندی API
    genai.configure(api_key="AIzaSyC0iaqJc8mw-D10aeORr05foUg0RudkAss")
    model = genai.GenerativeModel("gemini-1.5-flash")
    
    # ایجاد پرامپت
    prompt = (
        "برای محصول " + ws.cell(row=i, column=columns["Name"]).value +
        ' اطلاعات زیر را تولید کن: '
        '1. **توضیحات محصول**: یک پاراگراف شامل 5 جمله خاص و مرتبط با محصول که برای سئو بهینه باشد. '
        '2. **توضیح مختصر**: 1-2 جمله کوتاه و مفید درباره محصول. '
        '3. **کلمات کلیدی مرتبط**: حداقل 5 کلمه یا عبارت خاص محصول برای سئو. '
        '4. **برچسب‌ها (در صورت نیاز)**: اگر محصول مشابه دیگری در دسته مرتبط وجود دارد، برچسب مناسب ایجاد کن. در غیر این صورت، خالی بگذار.' 
        '**فرمت خروجی**:[[توضیحات محصول با 5 جمله],[توضیح مختصر در 1-2 جمله],[{"keyword":"کلمه کلیدی","score":امتیاز}, ...],[برچسب‌ها (در صورت نیاز)]] '
        'شما موظفید فقط و فقط فرمت خروجی را مطابق دستور زیر تولید کنید. هر گونه انحراف از فرمت مورد قبول نیست. '
        'حتی اگر نتیجه‌ای وجود نداشته باشد، قالب باید دقیقاً حفظ شود و تمام اجزا (مانند لیست‌ها یا کلیدهای JSON) به‌طور کامل وجود داشته باشند. '
        'اگر نتیجه‌ای وجود ندارد، مقدارها را به صورت خالی (مثلاً لیست خالی یا مقادیر null) تولید کنید. '
        'هیچ توضیح، متن اضافی، یا پیامی خارج از این فرمت نباید تولید شود.'
    )

    # حلقه برای تلاش مجدد در صورت ناموفق بودن
    attempts = 0
    while attempts < 3:
        try:
            # دریافت پاسخ از مدل
            response = model.generate_content(prompt)
            print("has given response...")
            res = response.text
            res = res.split('\n', 1)[1]  # حذف خطوط اضافی
            res = res.replace('```', '')
            res = res.strip()

            content = json.loads(res)  # تجزیه JSON
            break  # اگر موفق شدیم، از حلقه خارج می‌شویم

        except (json.JSONDecodeError, Exception) as e:
            print(f"Error on attempt {attempts + 1} for row {i}: {e}")
            attempts += 1
            if attempts < 3:
                print(f"Retrying for row {i}...")
                time.sleep(5)  # تأخیر برای جلوگیری از فشار زیاد به سرور
            else:
                attempts = 0

    # افزودن مقادیر به فایل اکسل
    try:
        ws.cell(row=i, column=columns["Product_Description"], value=content[0][0] if len(content[0]) > 0 else "")
        ws.cell(row=i, column=columns["Short_Description"], value=content[1][0] if len(content[1]) > 0 else "")
        
        # برای کلمات کلیدی، 'keyword' و 'score' باید به رشته تبدیل شوند
        seo_keywords = ", ".join([f'{item["keyword"]} ({item["score"]})' for item in content[2]]) if len(content[2]) > 0 else ""
        ws.cell(row=i, column=columns["SEO_Keywords"], value=seo_keywords)
        
        # برای برچسب‌ها، آنها را به رشته تبدیل می‌کنیم
        ws.cell(row=i, column=columns["Tags"], value=", ".join(content[3]) if len(content[3]) > 0 else "")
        
        # ذخیره تغییرات بعد از هر به‌روزرسانی
        wb.save(file_path)

    except ValueError as e:
        print(f"Error writing to Excel for row {i}: {e}")
        time.sleep(5)  # تأخیر برای جلوگیری از فشار زیاد به سرور یا فایل
        continue  # ادامه می‌دهیم و به ردیف بعدی می‌رویم

print("تمام ردیف‌ها پردازش شدند و فایل به‌روزرسانی شد.")
